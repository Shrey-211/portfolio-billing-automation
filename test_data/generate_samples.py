import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_sample_portfolio(filepath, client_data, holdings):
    wb = openpyxl.Workbook()
    
    # --- 1. holdings Sheet ---
    ws_holdings = wb.active
    ws_holdings.title = "Holdings"
    ws_holdings.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="1B365D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    double_bottom_border = Border(top=thin_side, bottom=Side(border_style="double", color="1B365D"))
    
    # Title Block
    ws_holdings["A1"] = f"PORTFOLIO PERFORMANCE REPORT"
    ws_holdings["A1"].font = font_title
    ws_holdings.row_dimensions[1].height = 25
    
    ws_holdings["A2"] = f"Client Name: {client_data['client_name']}"
    ws_holdings["A2"].font = Font(name="Calibri", size=11, italic=True)
    
    # Headers
    headers = ["Asset Category", "Security / Fund Name", "Quantity", "Purchase Price (INR)", "Current Price (INR)", "Current Value (INR)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_holdings.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws_holdings.row_dimensions[4].height = 24
    
    # Populate Holdings
    row_idx = 5
    for h in holdings:
        ws_holdings.cell(row=row_idx, column=1, value=h[0]).alignment = align_left
        ws_holdings.cell(row=row_idx, column=2, value=h[1]).alignment = align_left
        ws_holdings.cell(row=row_idx, column=3, value=h[2]).alignment = align_right
        ws_holdings.cell(row=row_idx, column=4, value=h[3]).alignment = align_right
        ws_holdings.cell(row=row_idx, column=5, value=h[4]).alignment = align_right
        
        # Formula for value: Quantity * Current Price
        # Coordinate for Quantity is C{row_idx}, Current Price is E{row_idx}
        formula_val = f"=C{row_idx}*E{row_idx}"
        ws_holdings.cell(row=row_idx, column=6, value=formula_val).alignment = align_right
        
        # Apply borders and formats
        for c in range(1, 7):
            cell = ws_holdings.cell(row=row_idx, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
            # Number formats
            if c in [4, 5]:
                cell.number_format = "#,##0.00"
            elif c == 6:
                cell.number_format = "#,##0.00"
                
        row_idx += 1
        
    # Total Valuation Row
    ws_holdings.cell(row=row_idx, column=1, value="Total Portfolio Value").font = font_bold
    ws_holdings.cell(row=row_idx, column=1).alignment = align_left
    ws_holdings.cell(row=row_idx, column=1).border = double_bottom_border
    
    # Blank columns
    for c in range(2, 6):
        ws_holdings.cell(row=row_idx, column=c).border = double_bottom_border
        
    # SUM Formula for column 6
    formula_sum = f"=SUM(F5:F{row_idx-1})"
    sum_cell = ws_holdings.cell(row=row_idx, column=6, value=formula_sum)
    sum_cell.font = font_bold
    sum_cell.alignment = align_right
    sum_cell.border = double_bottom_border
    sum_cell.number_format = "#,##0.00"
    
    # Adjust column dimensions
    for col in ws_holdings.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_holdings.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # --- 2. Metadata Sheet ---
    ws_meta = wb.create_sheet(title="Metadata")
    ws_meta.views.sheetView[0].showGridLines = True
    
    metadata_fields = [
        ("Client Name", client_data["client_name"]),
        ("Client Type", client_data["client_type"]),
        ("State", client_data["state"]),
        ("Address", client_data["address"]),
        ("GSTIN", client_data["gstin"]),
        ("Email", client_data["email"]),
        ("CC Email", client_data.get("cc_email", "")),
        ("Valuation", client_data["valuation"]) # Matches formula total value
    ]
    
    ws_meta.cell(row=1, column=1, value="Metadata Key").font = font_bold
    ws_meta.cell(row=1, column=2, value="Metadata Value").font = font_bold
    
    for r, (k, v) in enumerate(metadata_fields, 2):
        ws_meta.cell(row=r, column=1, value=k).font = font_bold
        val_cell = ws_meta.cell(row=r, column=2, value=v)
        val_cell.font = font_regular
        if k == "Valuation":
            val_cell.value = client_data["valuation"]
            
    ws_meta.column_dimensions['A'].width = 18
    ws_meta.column_dimensions['B'].width = 45
    
    # Save
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"Created sample portfolio: {filepath}")

def main():
    test_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_data")
    
    # Portfolio 1: Alpha Investors (MH, 1.25 Cr valuation)
    client_1 = {
        "client_name": "Alpha Investors LLC",
        "client_type": "Type 1",
        "state": "Maharashtra",
        "email": "alpha.investors@example.com",
        "cc_email": "advisor@example.com",
        "address": "123, BKC High Street, Bandra East, Mumbai, MH - 400051",
        "gstin": "27ALPHA1234A1Z0",
        "valuation": 12500000.0
    }
    holdings_1 = [
        ("Equity", "Reliance Industries Ltd", 2000, 2400.00, 2550.00),
        ("Equity", "HDFC Bank Ltd", 3000, 1600.00, 1580.00),
        ("Mutual Fund", "HDFC Top 100 Fund", 50000, 80.00, 92.50),
        ("Mutual Fund", "SBI Bluechip Fund", 25000, 50.00, 60.00),
        ("Debt", "ICICI Prudential Gilt Fund", 100000, 10.00, 10.80)
    ]
    
    # Portfolio 2: Beta Capital (Karnataka, 75 Lakhs valuation)
    client_2 = {
        "client_name": "Beta Capital",
        "client_type": "Type 2",
        "state": "Karnataka",
        "email": "beta.cap@example.com",
        "cc_email": "",
        "address": "456 Tech Corridor, Whitefield, Bengaluru, KA - 560066",
        "gstin": "29BETAC5678B1Z1",
        "valuation": 7500000.0
    }
    holdings_2 = [
        ("Equity", "Infosys Ltd", 1500, 1450.00, 1400.00),
        ("Equity", "Tata Consultancy Services", 500, 3200.00, 3400.00),
        ("Mutual Fund", "Mirae Asset Large Cap Fund", 20000, 120.00, 135.00),
        ("Mutual Fund", "Axis Midcap Fund", 10000, 80.00, 100.00),
    ]
    
    # Portfolio 3: Gamma Trust (MH, 6.5 Cr valuation - tests higher fee tier!)
    client_3 = {
        "client_name": "Gamma Trust Group",
        "client_type": "Type 3",
        "state": "Maharashtra",
        "email": "gamma.trust@example.com",
        "cc_email": "trustee@example.com",
        "address": "789 Marine Drive, Nariman Point, Mumbai, MH - 400021",
        "gstin": "27GAMMA9012C1Z2",
        "valuation": 65000000.0
    }
    holdings_3 = [
        ("Equity", "Larsen & Toubro Ltd", 5000, 2200.00, 2400.00),
        ("Equity", "Maruti Suzuki India Ltd", 1000, 8500.00, 9200.00),
        ("Equity", "Kotak Mahindra Bank Ltd", 8000, 1900.00, 1850.00),
        ("Mutual Fund", "Nippon India Small Cap Fund", 100000, 90.00, 115.00),
        ("Mutual Fund", "Parag Parikh Flexi Cap Fund", 50000, 50.00, 60.00),
        ("Debt", "HDFC Corporate Debt Fund", 1500000, 10.00, 10.60)
    ]
    
    create_sample_portfolio(os.path.join(test_dir, "alpha_investors.xlsx"), client_1, holdings_1)
    create_sample_portfolio(os.path.join(test_dir, "beta_capital.xlsx"), client_2, holdings_2)
    create_sample_portfolio(os.path.join(test_dir, "gamma_trust.xlsx"), client_3, holdings_3)
    
if __name__ == "__main__":
    main()
