import os
import sys
import threading
import time
from datetime import datetime
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import tkinter as tk
from tkinter import filedialog

import backend.database as database
import backend.processing as processing
import backend.email_service as email_service

app = FastAPI(title="Portfolio Billing API")

# Enable CORS for React development (port 5173)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global processing states
PROGRESS_STATE = {
    "running": False,
    "total": 0,
    "completed": 0,
    "failed": 0,
    "status": "Idle",
    "active_file": "",
    "logs": [],
    "batch_id": None
}

EMAIL_STATE = {
    "running": False,
    "total": 0,
    "completed": 0,
    "failed": 0,
    "logs": []
}

class SettingsUpdate(BaseModel):
    settings: Dict[str, str]

class FeeRule(BaseModel):
    min_value: float
    max_value: float
    percentage: float
    flat_rate: float = 0.0

class FeeRulesUpdate(BaseModel):
    rules: List[FeeRule]

class ClientUpdate(BaseModel):
    client_name: str
    client_type: str
    state: str
    email: str
    cc_email: str = ""
    address: str = ""
    gstin: str = ""

class FileItemToProcess(BaseModel):
    filename: str
    client_name: str
    client_type: str
    state: str
    email: str
    cc_email: str = ""
    address: str = ""
    gstin: str = ""
    valuation: float
    is_regular: bool = True
    custom_message: str = ""
    matched_sheet: Optional[str] = None
    fee_amount: Optional[float] = 0.0
    cgst: Optional[float] = 0.0
    sgst: Optional[float] = 0.0
    igst: Optional[float] = 0.0
    total_amount: Optional[float] = 0.0
    invoice_number: Optional[str] = ""
    rate: Optional[float] = 0.0
    particulars: Optional[str] = ""
    period_start: Optional[str] = ""
    period_end: Optional[str] = ""

class BatchProcessRequest(BaseModel):
    folder_path: str
    items: List[FileItemToProcess]
    import_mode: str = "folder"


class EmailSendRequest(BaseModel):
    item_ids: List[int]

def log_progress(msg: str, log_type: str = "info"):
    timestamp = datetime.now().strftime("%H:%M:%S")
    PROGRESS_STATE["logs"].append({
        "timestamp": timestamp,
        "message": msg,
        "type": log_type
    })

# --- SETTINGS ENDPOINTS ---
@app.get("/api/settings")
def get_settings():
    return database.get_all_settings()

@app.post("/api/settings")
def update_settings(payload: SettingsUpdate):
    try:
        database.save_all_settings(payload.settings)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/settings/fee-rules")
def get_fee_rules():
    return database.get_fee_rules()

@app.post("/api/settings/fee-rules")
def update_fee_rules(payload: FeeRulesUpdate):
    try:
        rules_dict = [r.model_dump() for r in payload.rules]
        database.save_fee_rules(rules_dict)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/clients")
def update_client_profile(payload: ClientUpdate):
    try:
        database.save_client(payload.model_dump())
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/clients/{client_name}")
def get_client_by_name(client_name: str):
    try:
        client = database.get_client(client_name)
        if not client:
            return {
                "client_name": client_name,
                "client_type": "Type 1",
                "state": "Maharashtra",
                "email": "",
                "cc_email": "",
                "address": "",
                "gstin": ""
            }
        return client
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- FOLDER DIALOG SELECT ---
@app.post("/api/folders/select")
def select_folder():
    # Native dialog runs in main thread context or separate process
    # Tkinter requires thread safety: run withdraw in a safe loop
    folder_path = ""
    try:
        # Run tkinter in a daemon-safe way
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        folder_path = filedialog.askdirectory(title="Select Portfolio Folder")
        root.destroy()
    except Exception as e:
        print(f"Error opening folder picker: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to open native picker: {e}")

    if not folder_path:
        return {"folder_path": "", "files": []}

    # Scan folder for excel files
    files = []
    try:
        for file in os.listdir(folder_path):
            if (file.endswith(".xlsx") or file.endswith(".xlsm")) and not file.startswith("~$"):
                fp = os.path.join(folder_path, file)
                
                # Fetch metadata
                meta = processing.get_client_metadata(fp)
                if not meta:
                    base = os.path.splitext(file)[0]
                    meta = {
                        "client_name": base.replace("_", " ").replace("-", " "),
                        "client_type": "Type 1",
                        "state": "Maharashtra",
                        "email": "",
                        "cc_email": "",
                        "address": "",
                        "gstin": "",
                        "valuation": 0.0
                    }
                
                # Cached DB check
                cached = database.get_client(meta["client_name"])
                if cached:
                    for k in ["client_type", "state", "email", "cc_email", "address", "gstin"]:
                        if not meta.get(k) and cached.get(k):
                            meta[k] = cached[k]
                            
                meta["filename"] = fp
                meta["status"] = "Pending"
                files.append(meta)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error scanning folder: {e}")

    return {"folder_path": os.path.normpath(folder_path), "files": files}

def apply_ui_overrides_to_excel(excel_path, item):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        master_sheet = processing.find_master_sheet_name(wb.sheetnames)
        if master_sheet:
            ws = wb[master_sheet]
            header_row, cols = processing.get_master_sheet_mapping(ws)
            if header_row is None:
                wb.close()
                return False
            
            # Find the row for this client
            target_row = None
            client_name_key = str(item["client_name"]).strip()
            
            # Find the client name column dynamically
            col_name_idx = processing.get_mapped_col(cols, ["client_name", "name_of_client"], 3)
            
            for r in range(header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=col_name_idx).value
                if cell_val and str(cell_val).strip().lower() == client_name_key.lower():
                    target_row = r
                    break
            
            if target_row:
                # Update cells using mapped indices
                # 1. Valuation
                col_val_idx = processing.get_mapped_col(cols, ["value", "valuation", "value_of_shares", "aum"])
                if col_val_idx:
                    ws.cell(row=target_row, column=col_val_idx).value = float(item["valuation"])
                
                # 2. Client Address
                col_addr_idx = processing.get_mapped_col(cols, ["address_1", "address"])
                if col_addr_idx:
                    ws.cell(row=target_row, column=col_addr_idx).value = item["address"]
                
                # 3. Client State
                state_val = item["state"]
                col_state_idx = processing.get_mapped_col(cols, ["state"])
                if col_state_idx:
                    ws.cell(row=target_row, column=col_state_idx).value = state_val
                
                # 4. Client Email
                col_email_idx = processing.get_mapped_col(cols, ["client_email_id", "email"])
                if col_email_idx:
                    ws.cell(row=target_row, column=col_email_idx).value = item["email"]
                
                # 5. Invoice No
                col_inv_idx = processing.get_mapped_col(cols, ["invoice_no", "invoice_number"])
                if col_inv_idx:
                    ws.cell(row=target_row, column=col_inv_idx).value = item["invoice_number"]
                
                # Update taxable amount and tax formulas dynamically based on settings & state
                from openpyxl.utils import get_column_letter
                is_maharashtra = str(state_val).strip().lower() == "maharashtra"
                
                settings_db = database.get_all_settings()
                formula_template = settings_db.get("taxable_amt_formula", "Value * Rate / 4")
                cgst_pct = float(settings_db.get("gst_rate_cgst", "9.0"))
                sgst_pct = float(settings_db.get("gst_rate_sgst", "9.0"))
                igst_pct = float(settings_db.get("gst_rate_igst", "18.0"))
                
                col_rate_idx = processing.get_mapped_col(cols, ["fee_@", "rate"])
                col_taxable_idx = processing.get_mapped_col(cols, ["taxable_amt", "taxable_amount"])
                
                # Write dynamic Taxable Amount formula
                if col_taxable_idx and col_val_idx and col_rate_idx:
                    val_let = get_column_letter(col_val_idx)
                    rate_let = get_column_letter(col_rate_idx)
                    
                    excel_formula = formula_template.replace("Value", f"{val_let}{target_row}").replace("Rate", f"{rate_let}{target_row}")
                    if not excel_formula.startswith("="):
                        excel_formula = "=" + excel_formula
                    
                    ws.cell(row=target_row, column=col_taxable_idx).value = excel_formula
                
                if col_taxable_idx:
                    tax_let = get_column_letter(col_taxable_idx)
                    
                    col_cgst_idx = processing.get_mapped_col(cols, ["cgst"])
                    col_sgst_idx = processing.get_mapped_col(cols, ["sgst"])
                    col_igst_idx = processing.get_mapped_col(cols, ["igst", "isgt"])
                    
                    if col_cgst_idx:
                        ws.cell(row=target_row, column=col_cgst_idx).value = f"={tax_let}{target_row}*{cgst_pct}%" if is_maharashtra else 0
                    if col_sgst_idx:
                        ws.cell(row=target_row, column=col_sgst_idx).value = f"={tax_let}{target_row}*{sgst_pct}%" if is_maharashtra else 0
                    if col_igst_idx:
                        ws.cell(row=target_row, column=col_igst_idx).value = 0 if is_maharashtra else f"={tax_let}{target_row}*{igst_pct}%"
                    
                    col_total_idx = processing.get_mapped_col(cols, ["total", "total_inv_amt", "total_amount"])
                    if col_total_idx:
                        cgst_let = get_column_letter(col_cgst_idx) if col_cgst_idx else ""
                        sgst_let = get_column_letter(col_sgst_idx) if col_sgst_idx else ""
                        igst_let = get_column_letter(col_igst_idx) if col_igst_idx else ""
                        
                        formula_parts = [f"{tax_let}{target_row}"]
                        if cgst_let: formula_parts.append(f"{cgst_let}{target_row}")
                        if sgst_let: formula_parts.append(f"{sgst_let}{target_row}")
                        if igst_let: formula_parts.append(f"{igst_let}{target_row}")
                        
                        ws.cell(row=target_row, column=col_total_idx).value = "=" + "+".join(formula_parts)
                
                wb.save(excel_path)
                wb.close()
                return True
        wb.close()
    except Exception as e:
        print(f"Error applying overrides to Excel: {e}")
    return False

def read_calculated_values_from_excel(excel_path, client_name):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        master_sheet = processing.find_master_sheet_name(wb.sheetnames)
        if master_sheet:
            ws = wb[master_sheet]
            header_row, cols = processing.get_master_sheet_mapping(ws)
            if header_row is None:
                wb.close()
                return None
            
            col_name_idx = processing.get_mapped_col(cols, ["client_name", "name_of_client"], 3)
            
            for r in range(header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=col_name_idx).value
                if cell_val and str(cell_val).strip().lower() == str(client_name).strip().lower():
                    col_taxable_idx = processing.get_mapped_col(cols, ["taxable_amt", "taxable_amount"], 17)
                    col_cgst_idx = processing.get_mapped_col(cols, ["cgst"], 18)
                    col_sgst_idx = processing.get_mapped_col(cols, ["sgst"], 19)
                    col_igst_idx = processing.get_mapped_col(cols, ["igst", "isgt"], 20)
                    col_total_idx = processing.get_mapped_col(cols, ["total", "total_inv_amt", "total_amount"], 21)
                    col_inv_idx = processing.get_mapped_col(cols, ["invoice_no", "invoice_number"], 5)
                    
                    res = {
                        "fee_amount": ws.cell(row=r, column=col_taxable_idx).value or 0.0,
                        "cgst": ws.cell(row=r, column=col_cgst_idx).value or 0.0,
                        "sgst": ws.cell(row=r, column=col_sgst_idx).value or 0.0,
                        "igst": ws.cell(row=r, column=col_igst_idx).value or 0.0,
                        "total_amount": ws.cell(row=r, column=col_total_idx).value or 0.0,
                        "invoice_number": ws.cell(row=r, column=col_inv_idx).value or ""
                    }
                    wb.close()
                    return res
        wb.close()
    except Exception as e:
        print(f"Error reading calculated values: {e}")
    return None

# --- BATCH RUN PROCESSOR ---
def run_batch_thread(folder_path: str, items: List[Dict[str, Any]], import_mode: str = "folder"):
    global PROGRESS_STATE
    total = len(items)
    completed = 0
    failed = 0
    
    PROGRESS_STATE["running"] = True
    PROGRESS_STATE["total"] = total
    PROGRESS_STATE["completed"] = 0
    PROGRESS_STATE["failed"] = 0
    PROGRESS_STATE["status"] = "Running"
    PROGRESS_STATE["logs"] = []
    
    log_progress("Starting batch billing automation run...", "info")
    
    # 1. Create Batch
    try:
        batch_id = database.create_batch(folder_path, total)
        PROGRESS_STATE["batch_id"] = batch_id
        log_progress(f"Created batch record #{batch_id} in SQLite database.", "info")
    except Exception as e:
        log_progress(f"DB Batch logging failed: {e}", "error")
        batch_id = 0
        
    pdf_dir = os.path.join(folder_path, "pdf")
    os.makedirs(pdf_dir, exist_ok=True)
    static_xlsx_dir = os.path.join(pdf_dir, "processed_xlsx")
    os.makedirs(static_xlsx_dir, exist_ok=True)
    
    settings = database.get_all_settings()
    fee_rules = database.get_fee_rules()
    
    fee_type = settings.get("fee_calculation_type", "flat")
    gst_rates = {
        "gst_rate_cgst": settings.get("gst_rate_cgst", "9.0"),
        "gst_rate_sgst": settings.get("gst_rate_sgst", "9.0"),
        "gst_rate_igst": settings.get("gst_rate_igst", "18.0"),
    }
    invoice_prefix = settings.get("invoice_prefix", "INV-2026-")
    try:
        next_inv_num = int(settings.get("next_invoice_number", "1"))
    except:
        next_inv_num = 1
        
    log_progress(f"Configurations loaded. Fee engine mode: {fee_type.upper()}.", "info")

    # Setup copies of Excel sheet if in sheet mode
    master_xlsx_copy = ""
    if import_mode == "sheet" and items:
        import shutil
        first_item_filename = items[0]["filename"]
        master_xlsx_copy = os.path.join(static_xlsx_dir, "Master_Processed.xlsx")
        try:
            shutil.copy2(first_item_filename, master_xlsx_copy)
            log_progress(f"Created working copy of master billing sheet: {os.path.basename(master_xlsx_copy)}", "info")
        except Exception as copy_err:
            log_progress(f"Failed to copy master sheet: {copy_err}", "error")
            master_xlsx_copy = first_item_filename
    
    for idx, item in enumerate(items):
        if not PROGRESS_STATE["running"]:
            log_progress("Batch run cancelled by user.", "warning")
            if batch_id > 0:
                database.update_batch(batch_id, completed, failed, "Cancelled")
            return
            
        filename_full = item["filename"]
        client_name = item["client_name"]
        filename_base = os.path.basename(filename_full)
        is_regular = item.get("is_regular", True)
        custom_message = item.get("custom_message", "")
        
        PROGRESS_STATE["active_file"] = f"{client_name} ({filename_base})"
        
        if import_mode == "sheet" and not is_regular:
            log_progress(f"[{idx+1}/{total}] Skipping non-billable client: {client_name}", "info")
            if batch_id > 0:
                try:
                    db_item = {
                        "batch_id": batch_id,
                        "filename": filename_base,
                        "client_name": client_name,
                        "valuation": item["valuation"],
                        "fee_amount": 0.0,
                        "cgst": 0.0,
                        "sgst": 0.0,
                        "igst": 0.0,
                        "total_amount": 0.0,
                        "status": "Skipped (Not Billable)",
                        "error_msg": "",
                        "portfolio_pdf_path": "",
                        "invoice_pdf_path": "",
                        "email_status": "Skipped",
                        "custom_message": custom_message
                    }
                    database.add_job_item(db_item)
                except Exception as db_ex:
                    log_progress(f"Database logging error: {db_ex}", "error")
            completed += 1
            PROGRESS_STATE["completed"] = completed
            continue
            
        log_progress(f"[{idx+1}/{total}] Processing client: {client_name}", "info")
        
        safe_client_name = "".join([c for c in client_name if c.isalnum() or c in (" ", "_", "-")]).strip().replace(" ", "_")
        flat_excel_path = os.path.join(static_xlsx_dir, f"{safe_client_name}_Processed.xlsx")
        
        matched_sheet = item.get("matched_sheet")
        portfolio_pdf_path = ""
        if import_mode == "sheet":
            if matched_sheet:
                portfolio_pdf_path = os.path.join(pdf_dir, f"{safe_client_name}_Portfolio.pdf")
        else:
            portfolio_pdf_path = os.path.join(pdf_dir, f"{safe_client_name}_Portfolio.pdf")
            
        invoice_pdf_path = os.path.join(pdf_dir, f"{safe_client_name}_Invoice.pdf")
        
        error_msg = ""
        success = False
        
        valuation = item["valuation"]
        fee_amount = 0.0
        cgst, sgst, igst, total_amount = 0.0, 0.0, 0.0, 0.0
        
        try:
            if import_mode == "sheet":
                log_progress(" Applying UI overrides to master sheet...", "debug")
                apply_ui_overrides_to_excel(master_xlsx_copy, item)
                
                # Check Excel availability
                if processing.EXCEL_AVAILABLE:
                    log_progress(" Generating tax invoice from Excel template...", "debug")
                    processing.generate_invoice_pdf_via_excel(master_xlsx_copy, client_name, invoice_pdf_path)
                    
                    if matched_sheet:
                        log_progress(f" Generating portfolio report from sheet: {matched_sheet}...", "debug")
                        processing.convert_excel_to_pdf(master_xlsx_copy, portfolio_pdf_path, matched_sheet)
                else:
                    # Fallback ReportLab Generation
                    log_progress(" Excel not available. Running fallback ReportLab generators...", "warning")
                    # Set up invoice details
                    invoice_data = {
                        "invoice_number": item.get("invoice_number") or f"{invoice_prefix}{next_inv_num:04d}",
                        "date": datetime.today().strftime('%d-%b-%Y'),
                        "client_name": client_name,
                        "client_address": item["address"],
                        "client_state": item["state"],
                        "client_gstin": item.get("gstin", ""),
                        "valuation": item["valuation"],
                        "fee_amount": item.get("fee_amount", 0.0),
                        "cgst": item.get("cgst", 0.0),
                        "sgst": item.get("sgst", 0.0),
                        "igst": item.get("igst", 0.0),
                        "total_amount": item.get("total_amount", 0.0),
                        "particulars": item.get("particulars", ""),
                        "rate": item.get("rate", 0.0),
                        "period_start": item.get("period_start", ""),
                        "period_end": item.get("period_end", "")
                    }
                    processing.generate_invoice_pdf(invoice_pdf_path, invoice_data, settings)
                    
                    if matched_sheet:
                        processing.convert_excel_to_pdf_fallback(master_xlsx_copy, portfolio_pdf_path, matched_sheet)

                # Now read calculated values from Excel copy (if excel was run, it recalculated)
                calc_vals = read_calculated_values_from_excel(master_xlsx_copy, client_name)
                if calc_vals:
                    fee_amount = calc_vals["fee_amount"]
                    cgst = calc_vals["cgst"]
                    sgst = calc_vals["sgst"]
                    igst = calc_vals["igst"]
                    total_amount = calc_vals["total_amount"]
                    invoice_code = calc_vals["invoice_number"] or item.get("invoice_number") or f"{invoice_prefix}{next_inv_num:04d}"
                else:
                    # Use fallback values from item
                    fee_amount = item.get("fee_amount", 0.0)
                    cgst = item.get("cgst", 0.0)
                    sgst = item.get("sgst", 0.0)
                    igst = item.get("igst", 0.0)
                    total_amount = item.get("total_amount", 0.0)
                    invoice_code = item.get("invoice_number") or f"{invoice_prefix}{next_inv_num:04d}"
            else:
                # Folder mode logic (original logic)
                log_progress(" Cleansing excel formulas...", "debug")
                processing.cleanse_formulas(filename_full, flat_excel_path)
                
                log_progress(" Converting portfolio worksheets to PDF...", "debug")
                processing.convert_excel_to_pdf(flat_excel_path, portfolio_pdf_path)
                
                # Fee
                fee_amount = processing.calculate_fees(valuation, fee_rules, fee_type)
                log_progress(f" Fee amount calculated: INR {fee_amount:,.2f}", "debug")
                
                # GST
                state = item["state"]
                gst_calc = processing.calculate_gst(fee_amount, state, gst_rates)
                cgst = gst_calc["cgst"]
                sgst = gst_calc["sgst"]
                igst = gst_calc["igst"]
                total_amount = gst_calc["total_amount"]
                
                # PDF Invoice
                log_progress(" Rendering tax invoice document...", "debug")
                invoice_code = f"{invoice_prefix}{next_inv_num:04d}"
                today_str = datetime.today().strftime('%d-%b-%Y')
                invoice_data = {
                    "invoice_number": invoice_code,
                    "date": today_str,
                    "client_name": client_name,
                    "client_address": item["address"],
                    "client_state": state,
                    "client_gstin": item["gstin"],
                    "valuation": valuation,
                    "fee_amount": fee_amount,
                    "cgst": cgst,
                    "sgst": sgst,
                    "igst": igst,
                    "total_amount": total_amount
                }
                processing.generate_invoice_pdf(invoice_pdf_path, invoice_data, settings)
                next_inv_num += 1
            
            success = True
            completed += 1
            log_progress(f"Successfully generated files for {client_name}.", "success")
        except Exception as ex:
            failed += 1
            success = False
            error_msg = str(ex)
            log_progress(f"Failed processing {client_name}: {error_msg}", "error")
            
        # Write job item
        if batch_id > 0:
            try:
                db_item = {
                    "batch_id": batch_id,
                    "filename": filename_base,
                    "client_name": client_name,
                    "valuation": valuation,
                    "fee_amount": fee_amount,
                    "cgst": cgst,
                    "sgst": sgst,
                    "igst": igst,
                    "total_amount": total_amount,
                    "status": "Completed" if success else "Failed",
                    "error_msg": error_msg if not success else "",
                    "portfolio_pdf_path": portfolio_pdf_path if success else "",
                    "invoice_pdf_path": invoice_pdf_path if success else "",
                    "email_status": "Pending",
                    "custom_message": custom_message,
                    "invoice_number": invoice_code if success else ""
                }
                database.add_job_item(db_item)
            except Exception as db_ex:
                log_progress(f"Database logging error: {db_ex}", "error")
                
        PROGRESS_STATE["completed"] = completed
        PROGRESS_STATE["failed"] = failed
        
    log_progress(f"Batch completed! Successful: {completed} | Failed: {failed}", "info")
    
    if batch_id > 0:
        try:
            database.save_setting("next_invoice_number", next_inv_num)
            batch_status = "Completed" if failed == 0 else ("Failed" if completed == 0 else "Partially Completed")
            database.update_batch(batch_id, completed, failed, batch_status)
        except Exception as db_ex:
            log_progress(f"DB finalize error: {db_ex}", "error")
            
    PROGRESS_STATE["running"] = False
    PROGRESS_STATE["status"] = "Completed"

@app.post("/api/batch/process")
def start_batch_processing(payload: BatchProcessRequest, background_tasks: BackgroundTasks):
    global PROGRESS_STATE
    if PROGRESS_STATE["running"]:
        raise HTTPException(status_code=400, detail="A batch job is already running.")
        
    items_dict = [item.model_dump() for item in payload.items]
    PROGRESS_STATE = {
        "running": True,
        "total": len(items_dict),
        "completed": 0,
        "failed": 0,
        "status": "Starting",
        "active_file": "",
        "logs": [],
        "batch_id": None
    }
    background_tasks.add_task(run_batch_thread, payload.folder_path, items_dict, payload.import_mode)
    return {"status": "started"}

@app.get("/api/batch/progress")
def get_batch_progress():
    return PROGRESS_STATE

@app.post("/api/batch/cancel")
def cancel_batch_processing():
    global PROGRESS_STATE
    if PROGRESS_STATE["running"]:
        PROGRESS_STATE["running"] = False
        return {"status": "cancelling"}
    return {"status": "not running"}

# --- RESULTS HISTORY ---
@app.get("/api/batch/recent")
def get_recent_batches():
    return database.get_recent_batches(10)

@app.get("/api/batch/{batch_id}/results")
def get_batch_results(batch_id: int):
    return database.get_batch_items(batch_id)

@app.post("/api/pdf/open")
def open_pdf_document(payload: Dict[str, str]):
    path = payload.get("path")
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    try:
        os.startfile(path)
        return {"status": "opened"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/folders/open")
def open_folder_directory(payload: Dict[str, str]):
    path = payload.get("path")
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Directory not found")
    try:
        os.startfile(path)
        return {"status": "opened"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- BULK EMAIL DELIVER ---
def run_email_thread(item_ids: List[int]):
    global EMAIL_STATE
    EMAIL_STATE["running"] = True
    EMAIL_STATE["logs"] = []
    EMAIL_STATE["total"] = len(item_ids)
    EMAIL_STATE["completed"] = 0
    EMAIL_STATE["failed"] = 0
    
    settings = database.get_all_settings()
    subject_template = settings.get("email_subject_template", "Portfolio Report & Invoice")
    body_template = settings.get("email_body_template", "Dear {ClientName},\n...")
    
    # Retrieve job items
    # For simplicity, we can fetch items from DB using SQL queries
    conn = database.get_db_connection()
    cursor = conn.cursor()
    
    for idx, item_id in enumerate(item_ids):
        if not EMAIL_STATE["running"]:
            EMAIL_STATE["logs"].append("Delivery cancelled by user.")
            break
            
        cursor.execute("SELECT * FROM job_items WHERE id = ?", (item_id,))
        row = cursor.fetchone()
        if not row:
            EMAIL_STATE["failed"] += 1
            EMAIL_STATE["logs"].append(f"Item #{item_id} not found in DB.")
            continue
            
        item = dict(row)
        client_name = item["client_name"]
        
        # Load registry detail
        client_info = database.get_client(client_name)
        if not client_info:
            client_info = {
                "client_name": client_name,
                "email": item.get("email", ""),
                "cc_email": item.get("cc_email", ""),
            }
            
        to_email = client_info.get("email")
        if not to_email:
            EMAIL_STATE["failed"] += 1
            EMAIL_STATE["logs"].append(f"Failed {client_name}: Email missing.")
            database.update_job_item(item_id, {"email_status": "Failed: No Email"})
            EMAIL_STATE["completed"] += 1
            continue
            
        EMAIL_STATE["logs"].append(f"Delivering email to {client_name} ({to_email})...")
        
        attachments = []
        if item["portfolio_pdf_path"] and os.path.exists(item["portfolio_pdf_path"]):
            attachments.append(item["portfolio_pdf_path"])
        if item["invoice_pdf_path"] and os.path.exists(item["invoice_pdf_path"]):
            attachments.append(item["invoice_pdf_path"])
            
        if not attachments:
            EMAIL_STATE["failed"] += 1
            EMAIL_STATE["logs"].append(f"Failed {client_name}: Documents missing.")
            database.update_job_item(item_id, {"email_status": "Failed: Missing Docs"})
            EMAIL_STATE["completed"] += 1
            continue
            
        invoice_code = os.path.basename(item["invoice_pdf_path"]).replace("_Invoice.pdf", "") if item["invoice_pdf_path"] else "N/A"
        invoice_details = {
            "invoice_number": invoice_code,
            "valuation": item["valuation"],
            "fee_amount": item["fee_amount"],
            "total_amount": item["total_amount"],
            "custom_message": item.get("custom_message", "")
        }
        
        try:
            email_service.send_client_email(
                settings,
                client_info,
                subject_template,
                body_template,
                invoice_details,
                attachments,
                display_outlook=False
            )
            EMAIL_STATE["completed"] += 1
            EMAIL_STATE["logs"].append(f"Success: Sent reports to {client_name}.")
            database.update_job_item(item_id, {"email_status": "Sent"})
        except Exception as e:
            EMAIL_STATE["failed"] += 1
            EMAIL_STATE["logs"].append(f"Failed {client_name}: {e}")
            database.update_job_item(item_id, {"email_status": f"Failed: {str(e)[:40]}"})
            
    conn.close()
    EMAIL_STATE["running"] = False

@app.post("/api/email/send")
def send_bulk_emails(payload: EmailSendRequest, background_tasks: BackgroundTasks):
    global EMAIL_STATE
    if EMAIL_STATE["running"]:
        raise HTTPException(status_code=400, detail="Email sender is busy.")
    background_tasks.add_task(run_email_thread, payload.item_ids)
    return {"status": "started"}

@app.get("/api/email/progress")
def get_email_progress():
    return EMAIL_STATE

# --- EXCEL SINGLE SHEET PARSE & LEDGER ENDPOINTS ---
@app.post("/api/excel/select")
def select_excel_file():
    excel_path = ""
    try:
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        excel_path = filedialog.askopenfilename(
            title="Select Billing Excel Sheet",
            filetypes=[("Excel Files", "*.xlsx *.xlsm")]
        )
        root.destroy()
    except Exception as e:
        print(f"Error opening file picker: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to open native file picker: {e}")

    if not excel_path:
        return {"folder_path": "", "files": []}

    try:
        company_details, clients = processing.parse_single_excel_sheet(excel_path)
        if company_details:
            database.save_all_settings(company_details)
        # Apply cached registry check for client info if details are missing
        for client in clients:
            cached = database.get_client(client["client_name"])
            if cached:
                for k in ["client_type", "state", "email", "cc_email", "address", "gstin"]:
                    if not client.get(k) and cached.get(k):
                        client[k] = cached[k]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error parsing Excel file: {e}")

    folder_path = os.path.dirname(excel_path)
    return {"folder_path": os.path.normpath(folder_path), "excel_path": excel_path, "files": clients}

@app.get("/api/invoices/all")
def get_all_invoices():
    try:
        return database.get_all_job_items(1000)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class PaymentUpdatePayload(BaseModel):
    is_paid: int
    payment_date: Optional[str] = None

@app.post("/api/invoices/{item_id}/payment")
def update_payment(item_id: int, payload: PaymentUpdatePayload):
    try:
        database.update_payment_status(item_id, payload.is_paid, payload.payment_date)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class JobItemEditPayload(BaseModel):
    client_name: str
    client_type: str
    state: str
    email: str
    cc_email: str = ""
    address: str = ""
    gstin: str = ""
    valuation: float
    is_regular: bool = True
    custom_message: str = ""

@app.post("/api/invoices/{item_id}/edit")
def edit_job_item(item_id: int, payload: JobItemEditPayload):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM job_items WHERE id = ?", (item_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Invoice record not found")
            
        item = dict(row)
        batch_id = item["batch_id"]
        
        # Get batch folder_path
        cursor.execute("SELECT folder_path FROM batches WHERE id = ?", (batch_id,))
        batch_row = cursor.fetchone()
        folder_path = batch_row["folder_path"] if batch_row else "."
        conn.close()
        
        # Recalculate fee and tax
        valuation = payload.valuation
        fee_rules = database.get_fee_rules()
        settings = database.get_all_settings()
        fee_type = settings.get("fee_calculation_type", "flat")
        gst_rates = {
            "gst_rate_cgst": settings.get("gst_rate_cgst", "9.0"),
            "gst_rate_sgst": settings.get("gst_rate_sgst", "9.0"),
            "gst_rate_igst": settings.get("gst_rate_igst", "18.0"),
        }
        
        is_regular = payload.is_regular
        if not is_regular:
            fee_amount, cgst, sgst, igst, total_amount = 0.0, 0.0, 0.0, 0.0, 0.0
            status = "Skipped (Not Billable)"
            invoice_pdf_path = ""
        else:
            fee_amount = processing.calculate_fees(valuation, fee_rules, fee_type)
            gst_calc = processing.calculate_gst(fee_amount, payload.state, gst_rates)
            cgst = gst_calc["cgst"]
            sgst = gst_calc["sgst"]
            igst = gst_calc["igst"]
            total_amount = gst_calc["total_amount"]
            status = "Completed"
            
            # Paths
            pdf_dir = os.path.join(folder_path, "pdf")
            os.makedirs(pdf_dir, exist_ok=True)
            safe_client_name = "".join([c for c in payload.client_name if c.isalnum() or c in (" ", "_", "-")]).strip().replace(" ", "_")
            invoice_pdf_path = os.path.join(pdf_dir, f"{safe_client_name}_Invoice.pdf")
            
            # Generate invoice number if not existing
            invoice_code = item.get("invoice_number")
            if not invoice_code:
                invoice_prefix = settings.get("invoice_prefix", "INV-2026-")
                invoice_code = f"{invoice_prefix}{item_id:04d}"
                
            # Render PDF Invoice
            invoice_data = {
                "invoice_number": invoice_code,
                "date": datetime.today().strftime('%d-%b-%Y'),
                "client_name": payload.client_name,
                "client_address": payload.address,
                "client_state": payload.state,
                "client_gstin": payload.gstin,
                "valuation": valuation,
                "fee_amount": fee_amount,
                "cgst": cgst,
                "sgst": sgst,
                "igst": igst,
                "total_amount": total_amount
            }
            processing.generate_invoice_pdf(invoice_pdf_path, invoice_data, settings)
            
        # Update client profile in registry
        database.save_client({
            "client_name": payload.client_name,
            "client_type": payload.client_type,
            "state": payload.state,
            "email": payload.email,
            "cc_email": payload.cc_email,
            "address": payload.address,
            "gstin": payload.gstin
        })
        
        # Update database record
        updates = {
            "client_name": payload.client_name,
            "valuation": valuation,
            "fee_amount": fee_amount,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "total_amount": total_amount,
            "status": status,
            "invoice_pdf_path": invoice_pdf_path,
            "custom_message": payload.custom_message,
            "email_status": "Pending" if is_regular else "Skipped"
        }
        database.update_job_item(item_id, updates)
        
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/invoices/export")
def export_billing_ledger(payload: Dict[str, str]):
    folder_path = payload.get("folder_path")
    if not folder_path or not os.path.exists(folder_path):
        raise HTTPException(status_code=400, detail="Invalid output folder path")
        
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM job_items ORDER BY id DESC")
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        if not rows:
            raise HTTPException(status_code=400, detail="No billing data to export")
            
        import pandas as pd
        data = []
        for r in rows:
            data.append({
                "Invoice ID": r["id"],
                "Client Name": r["client_name"],
                "Valuation (INR)": r["valuation"],
                "Fee Amount (INR)": r["fee_amount"],
                "CGST (INR)": r["cgst"],
                "SGST (INR)": r["sgst"],
                "IGST (INR)": r["igst"],
                "Total Invoice Value (INR)": r["total_amount"],
                "Generation Status": r["status"],
                "Email Status": r["email_status"],
                "Payment Status": "Paid" if r["is_paid"] == 1 else "Unpaid",
                "Payment Received Date": r["payment_date"] or ""
            })
            
        df = pd.DataFrame(data)
        
        pdf_dir = os.path.join(folder_path, "pdf")
        os.makedirs(pdf_dir, exist_ok=True)
        export_path = os.path.join(pdf_dir, "Master_Billing_Ledger.xlsx")
        
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Billing Ledger', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Billing Ledger']
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='CBD5E0'),
                right=Side(style='thin', color='CBD5E0'),
                top=Side(style='thin', color='CBD5E0'),
                bottom=Side(style='thin', color='CBD5E0')
            )
            
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    cell.border = thin_border
                    val_str = str(cell.value or '')
                    max_len = max(max_len, len(val_str))
                    
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = align_center
                    else:
                        if cell.column in [3, 4, 5, 6, 7, 8]:
                            cell.alignment = align_right
                            if cell.value is not None:
                                cell.number_format = '₹#,##0.00'
                        elif cell.column in [1, 9, 10, 11, 12]:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left
                            
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        return {"status": "success", "file_path": os.path.normpath(export_path)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# --- STATIC FILES MOUNT ---
# In production, we mount the frontend/dist directory to serve React SPA assets.
base_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
dist_dir = os.path.normpath(os.path.join(base_dir, "frontend", "dist"))

# Mount if the directory exists (it will exist after npm run build)
if os.path.exists(dist_dir):
    app.mount("/", StaticFiles(directory=dist_dir, html=True), name="static")
else:
    print(f"Warning: React static distribution folder not found at {dist_dir}. Serving API only.")

