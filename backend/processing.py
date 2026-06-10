import os
import sys
import openpyxl
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# For Windows Excel Automation
try:
    import win32com.client
    import pythoncom
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def get_client_metadata(excel_path):
    """
    Scans the excel file for a sheet named 'Metadata' or 'ClientInfo' (case-insensitive).
    Extracts key-value pairs from Column A and B.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading {excel_path} for metadata: {e}")
        return None
        
    metadata_sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in ["metadata", "clientinfo", "client_info"]:
            metadata_sheet = name
            break
            
    if not metadata_sheet:
        wb.close()
        return None
        
    sheet = wb[metadata_sheet]
    metadata = {}
    
    # Read rows (up to 50 rows for safety)
    for r in range(1, 50):
        key_cell = sheet.cell(row=r, column=1).value
        val_cell = sheet.cell(row=r, column=2).value
        
        if key_cell is not None:
            key = str(key_cell).strip().lower().replace(" ", "_")
            metadata[key] = val_cell
            
    wb.close()
    
    # Map to standard keys
    mapped = {
        "client_name": metadata.get("client_name") or metadata.get("name") or "",
        "client_type": metadata.get("client_type") or metadata.get("type") or "Type 1",
        "state": metadata.get("state") or metadata.get("client_state") or "Maharashtra",
        "email": metadata.get("email") or metadata.get("client_email") or "",
        "cc_email": metadata.get("cc_email") or metadata.get("cc") or "",
        "address": metadata.get("address") or metadata.get("client_address") or "",
        "gstin": metadata.get("gstin") or metadata.get("client_gstin") or "",
    }
    
    # Try to parse valuation
    val_raw = metadata.get("valuation") or metadata.get("portfolio_valuation") or metadata.get("portfolio_value") or 0.0
    try:
        # Strip currency symbols and commas if it's a string
        if isinstance(val_raw, str):
            val_raw = val_raw.replace("₹", "").replace("$", "").replace(",", "").strip()
        mapped["valuation"] = float(val_raw)
    except Exception:
        mapped["valuation"] = 0.0
        
    return mapped

def cleanse_formulas(input_path, output_path):
    """
    Reads excel from input_path, converts all formulas to static values
    using the cached values from Excel, preserving all formatting and layout,
    and writes to output_path.
    """
    try:
        # Load workbook with data_only=True to get evaluated values
        wb_data = openpyxl.load_workbook(input_path, data_only=True)
        # Load workbook with data_only=False to keep formatting/properties
        wb_styled = openpyxl.load_workbook(input_path, data_only=False)
        
        for sheet_name in wb_styled.sheetnames:
            sheet_styled = wb_styled[sheet_name]
            # Some sheets might be missing in data-only mode due to openpyxl issues, check first
            if sheet_name not in wb_data.sheetnames:
                continue
            sheet_data = wb_data[sheet_name]
            
            # Iterate through all cells in the styled sheet
            for r in range(1, sheet_styled.max_row + 1):
                for c in range(1, sheet_styled.max_column + 1):
                    cell_styled = sheet_styled.cell(row=r, column=c)
                    val = cell_styled.value
                    
                    # If it's a formula (string starting with '=')
                    if isinstance(val, str) and val.startswith("="):
                        # Copy the evaluated value from the data sheet
                        eval_val = sheet_data.cell(row=r, column=c).value
                        cell_styled.value = eval_val
                        
        # Ensure output folder exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb_styled.save(output_path)
        wb_data.close()
        wb_styled.close()
        return True
    except Exception as e:
        print(f"Error in formula cleansing: {e}")
        import traceback
        traceback.print_exc()
        raise e

def calculate_fees(valuation, rules, calculation_type="flat"):
    """
    Calculates fees based on rules.
    rules: list of dicts: {'min_value': X, 'max_value': Y, 'percentage': P, 'flat_rate': F}
    calculation_type: 'flat' or 'slab'
    """
    fee = 0.0
    if calculation_type == "flat":
        # Find the matching bracket
        for rule in rules:
            if rule["min_value"] <= valuation < rule["max_value"]:
                fee = (valuation * rule["percentage"] / 100.0) + rule["flat_rate"]
                break
        else:
            # Fallback to the last rule if value exceeds all
            if rules:
                rule = rules[-1]
                fee = (valuation * rule["percentage"] / 100.0) + rule["flat_rate"]
    else:
        # Slab-based progressive calculation
        remaining = valuation
        for rule in rules:
            min_val = rule["min_value"]
            max_val = rule["max_value"]
            pct = rule["percentage"]
            flat = rule["flat_rate"]
            
            if valuation > min_val:
                taxable_in_slab = min(valuation, max_val) - min_val
                if taxable_in_slab > 0:
                    fee += (taxable_in_slab * pct / 100.0) + flat
                    
    return fee

def calculate_gst(fee_amount, state, gst_rates):
    """
    Calculates GST based on client's state.
    gst_rates: dict with 'cgst', 'sgst', 'igst' float rates
    """
    is_maharashtra = str(state).strip().lower() == "maharashtra"
    
    cgst_pct = float(gst_rates.get("gst_rate_cgst", 9.0))
    sgst_pct = float(gst_rates.get("gst_rate_sgst", 9.0))
    igst_pct = float(gst_rates.get("gst_rate_igst", 18.0))
    
    if is_maharashtra:
        cgst = fee_amount * (cgst_pct / 100.0)
        sgst = fee_amount * (sgst_pct / 100.0)
        igst = 0.0
    else:
        cgst = 0.0
        sgst = 0.0
        igst = fee_amount * (igst_pct / 100.0)
        
    total_amount = fee_amount + cgst + sgst + igst
    return {
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "total_amount": total_amount,
        "is_maharashtra": is_maharashtra
    }

def convert_excel_to_pdf_fallback(excel_path, pdf_path, sheet_name=None):
    """
    Fallback method to generate a clean tabular report PDF from the Excel file
    using ReportLab when Microsoft Excel is not installed/configured.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        elif "Holdings" in wb.sheetnames:
            sheet = wb["Holdings"]
        else:
            sheet = wb.active
            
        rows = []
        # Scan from row 4 (where headers usually start in our sample template)
        for r in range(4, sheet.max_row + 1):
            row_vals = []
            has_val = False
            for c in range(1, 7):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    has_val = True
                row_vals.append(val)
            if has_val:
                rows.append(row_vals)
                
        wb.close()
        
        # Ensure pdf directory exists
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=45,
            bottomMargin=45
        )
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'PortfolioTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#1a365d')
        )
        
        normal_text = ParagraphStyle(
            'NormText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#4a5568')
        )
        
        bold_text = ParagraphStyle(
            'BoldText',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#2d3748')
        )
        
        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=12,
            textColor=colors.white
        )
        
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#2d3748')
        )
        
        table_cell_right = ParagraphStyle(
            'TableCellRight',
            parent=table_cell_style,
            alignment=2
        )
        
        table_cell_right_bold = ParagraphStyle(
            'TableCellRightBold',
            parent=table_cell_style,
            fontName='Helvetica-Bold',
            alignment=2
        )
        
        story = []
        story.append(Paragraph("PORTFOLIO PERFORMANCE REPORT", title_style))
        client_name_info = os.path.basename(pdf_path).replace("_Portfolio.pdf", "").replace("_", " ")
        story.append(Paragraph(f"<b>Client Profile:</b> {client_name_info} | Generated via Local Engine Fallback", normal_text))
        story.append(Spacer(1, 15))
        
        table_data = []
        if rows:
            # Header
            headers = [Paragraph(f"<b>{str(h)}</b>", table_header_style) for h in rows[0]]
            table_data.append(headers)
            
            # Data rows
            for r_idx, r_val in enumerate(rows[1:], 1):
                row_items = []
                is_last_row = (r_idx == len(rows) - 1)
                
                for c_idx, val in enumerate(r_val):
                    if val is None:
                        row_items.append("")
                    else:
                        if c_idx in [2, 3, 4, 5]:
                            try:
                                f_val = float(val)
                                formatted = f"{f_val:,.2f}" if c_idx != 2 or not f_val.is_integer() else f"{int(f_val)}"
                                style = table_cell_right_bold if is_last_row else table_cell_right
                                row_items.append(Paragraph(formatted, style))
                            except (ValueError, TypeError):
                                if hasattr(val, "strftime"):
                                    formatted = val.strftime('%d-%b-%Y')
                                else:
                                    formatted = str(val)
                                style = bold_text if is_last_row else table_cell_style
                                row_items.append(Paragraph(formatted, style))
                        else:
                            style = bold_text if is_last_row else table_cell_style
                            row_items.append(Paragraph(str(val), style))
                table_data.append(row_items)
                
        # Total A4 printable width is ~515 pt
        col_widths = [75, 170, 50, 70, 70, 80]
        t = Table(table_data, colWidths=col_widths)
        
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#cbd5e0')),
            ('PADDING', (0,0), (-1,-1), 6),
        ]
        if len(table_data) > 1:
            t_style.append(('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#1a365d')))
            t_style.append(('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.HexColor('#1a365d')))
            
        t.setStyle(TableStyle(t_style))
        story.append(t)
        doc.build(story)
        return True
    except Exception as e:
        print(f"Fallback portfolio PDF generation failed: {e}")
        raise e

def convert_excel_to_pdf(excel_path, pdf_path, sheet_name=None):
    """
    Uses win32com to open the excel file and export to PDF.
    Must be run in main thread or call pythoncom.CoInitialize() first if inside QThread.
    Falls back to ReportLab layout generation if Excel is unavailable.
    """
    if not EXCEL_AVAILABLE:
        print("win32com not available. Running ReportLab fallback for Portfolio PDF.")
        return convert_excel_to_pdf_fallback(excel_path, pdf_path, sheet_name)
        
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_excel = os.path.abspath(excel_path)
        abs_pdf = os.path.abspath(pdf_path)
        
        os.makedirs(os.path.dirname(abs_pdf), exist_ok=True)
        wb = excel.Workbooks.Open(abs_excel)
        if sheet_name and sheet_name in [s.Name for s in wb.Sheets]:
            ws = wb.Sheets(sheet_name)
            orig_visibility = ws.Visible
            if orig_visibility != -1:  # -1 represents xlSheetVisible
                ws.Visible = -1  # Temporarily make it visible
            try:
                ws.ExportAsFixedFormat(0, abs_pdf)
            finally:
                if orig_visibility != -1:
                    ws.Visible = orig_visibility  # Restore original visibility
        else:
            wb.ExportAsFixedFormat(0, abs_pdf)
        wb.Close(False)
        return True
    except Exception as e:
        print(f"Excel COM PDF conversion failed: {e}. Attempting ReportLab fallback.")
        try:
            return convert_excel_to_pdf_fallback(excel_path, pdf_path, sheet_name)
        except Exception as fallback_err:
            print(f"Both Excel COM and ReportLab fallback failed: {fallback_err}")
            raise e
    finally:
        if wb:
            try:
                wb.Close(False)
            except:
                pass
        if excel:
            try:
                excel.Quit()
            except:
                pass
        pythoncom.CoUninitialize()

def generate_invoice_pdf(pdf_path, invoice_details, company_details):
    """
    Generates a beautiful, professional, GST-compliant invoice PDF using reportlab.
    invoice_details: dict containing invoice_number, date, client_name, client_address,
                     client_gstin, valuation, fee_amount, cgst, sgst, igst, total_amount,
                     particulars, rate, period_start, period_end, etc.
    company_details: dict containing company_name, company_address, company_gstin,
                     company_bank_name, company_bank_account, company_bank_ifsc, company_bank_branch, etc.
    """
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=45,
        bottomMargin=45
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles for Premium Look
    title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#1a365d') # Dark Navy Blue
    )
    
    company_name_style = ParagraphStyle(
        'CompName',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=15,
        textColor=colors.HexColor('#2d3748')
    )
    
    normal_text = ParagraphStyle(
        'NormText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#4a5568')
    )
    
    bold_text = ParagraphStyle(
        'BoldText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#2d3748')
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#2d3748')
    )
    
    table_cell_right = ParagraphStyle(
        'TableCellRight',
        parent=table_cell_style,
        alignment=2 # Right align
    )

    table_cell_right_bold = ParagraphStyle(
        'TableCellRightBold',
        parent=table_cell_style,
        fontName='Helvetica-Bold',
        alignment=2 # Right align
    )

    story = []
    
    # --- HEADER SECTION ---
    # Two column header: Company details left, INVOICE title & metadata right
    company_p = Paragraph(
        f"<b>{company_details.get('company_name', 'Cedrus Consultants Pvt Ltd')}</b><br/>"
        f"{company_details.get('company_division', 'Investment Advisory Division')}<br/>"
        f"{company_details.get('company_address', '')}<br/>"
        f"GSTIN: {company_details.get('company_gstin', '')} | PAN: {company_details.get('company_pan', '')}<br/>"
        f"Email: {company_details.get('company_email', '')} | Phone: {company_details.get('company_phone', '')}",
        normal_text
    )
    
    invoice_meta_p = Paragraph(
        f"<font color='#1a365d'><b>TAX INVOICE</b></font><br/><br/>"
        f"<b>Invoice No:</b> {invoice_details.get('invoice_number', '')}<br/>"
        f"<b>Date:</b> {invoice_details.get('date', '')}<br/>"
        f"<b>Place of Supply:</b> {invoice_details.get('client_state', 'Maharashtra')}",
        normal_text
    )
    
    header_table = Table([[company_p, invoice_meta_p]], colWidths=[310, 205])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    
    story.append(header_table)
    story.append(Spacer(1, 15))
    
    # Decorative line
    divider = Table([['']], colWidths=[515])
    divider.setStyle(TableStyle([
        ('LINEABOVE', (0,0), (-1,-1), 1.5, colors.HexColor('#1a365d')),
        ('PADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(divider)
    story.append(Spacer(1, 15))
    
    # --- BILL TO SECTION ---
    bill_to_p = Paragraph(
        f"<b>BILL TO:</b><br/>"
        f"<b>{invoice_details.get('client_name', 'Client Name')}</b><br/>"
        f"Address: {invoice_details.get('client_address', 'N/A')}<br/>"
        f"State: {invoice_details.get('client_state', 'N/A')}<br/>"
        f"GSTIN: {invoice_details.get('client_gstin', 'N/A')}",
        normal_text
    )
    
    bill_table = Table([[bill_to_p]], colWidths=[515])
    bill_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f7fafc')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(bill_table)
    story.append(Spacer(1, 20))
    
    # --- PARTICULARS TABLE ---
    hsn_sac = str(invoice_details.get("sac") or invoice_details.get("sac_code") or "997159")
    
    particulars_text = invoice_details.get("particulars") or "Portfolio Management & Advisory Fees"
    # If valuation is present and not already in text, append it
    if "valuation" in invoice_details and str(invoice_details["valuation"]) not in particulars_text and float(invoice_details.get("valuation", 0.0)) > 0:
        particulars_text += f"<br/><font size='8' color='#718096'>Calculated on Portfolio Valuation of INR {invoice_details.get('valuation', 0.0):,.2f}</font>"
        
    particulars_p = Paragraph(particulars_text, table_cell_style)
    
    period_text = ""
    if invoice_details.get("period_start") or invoice_details.get("period_end"):
        p_start = invoice_details.get("period_start", "")
        p_end = invoice_details.get("period_end", "")
        if "00:00:00" in p_end:
            p_end = p_end.replace("00:00:00", "").strip()
        period_text = f"{p_start} to {p_end}" if p_start and p_end else (p_start or p_end)

    rate_val = invoice_details.get("rate", 0.0)
    if isinstance(rate_val, (int, float)):
        if rate_val < 1.0 and rate_val > 0:
            rate_text = f"{rate_val * 100:.4g}%"
        elif rate_val >= 1.0:
            rate_text = f"{rate_val:.4g}%"
        else:
            rate_text = "-"
    else:
        rate_text = str(rate_val)

    cols = ["Description", "SAC Code", "Period", "Rate", "Amount (INR)"]
    row_data = [
        [Paragraph(f"<b>{c}</b>", table_header_style) for c in cols],
        [
            particulars_p,
            Paragraph(hsn_sac, table_cell_style),
            Paragraph(period_text, table_cell_style),
            Paragraph(rate_text, table_cell_right),
            Paragraph(f"{invoice_details.get('fee_amount', 0.0):,.2f}", table_cell_right)
        ]
    ]
    
    # Add tax rows (CGST, SGST, IGST)
    cgst = invoice_details.get("cgst", 0.0)
    sgst = invoice_details.get("sgst", 0.0)
    igst = invoice_details.get("igst", 0.0)
    
    if cgst > 0 or sgst > 0:
        row_data.append([
            Paragraph("CGST @ 9%", table_cell_style), "", "", "",
            Paragraph(f"{cgst:,.2f}", table_cell_right)
        ])
        row_data.append([
            Paragraph("SGST @ 9%", table_cell_style), "", "", "",
            Paragraph(f"{sgst:,.2f}", table_cell_right)
        ])
    else:
        row_data.append([
            Paragraph("IGST @ 18%", table_cell_style), "", "", "",
            Paragraph(f"{igst:,.2f}", table_cell_right)
        ])
        
    # Total row
    row_data.append([
        Paragraph("<b>Total Invoice Value (INR)</b>", table_cell_style), "", "", "",
        Paragraph(f"<b>{invoice_details.get('total_amount', 0.0):,.2f}</b>", table_cell_right_bold)
    ])
    
    particulars_table = Table(row_data, colWidths=[200, 60, 110, 50, 95])
    
    # Style logic depending on tax rows count
    t_styles = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')),
        ('ALIGN', (0,0), (-1,0), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#cbd5e0')),
        ('LINEBELOW', (0,-1), (-1,-1), 1.5, colors.HexColor('#1a365d')),
        ('PADDING', (0,0), (-1,-1), 8),
    ]
    
    if cgst > 0 or sgst > 0:
        t_styles.append(('SPAN', (0,2), (3,2)))
        t_styles.append(('SPAN', (0,3), (3,3)))
        t_styles.append(('SPAN', (0,4), (3,4)))
    else:
        t_styles.append(('SPAN', (0,2), (3,2)))
        t_styles.append(('SPAN', (0,3), (3,3)))
        
    particulars_table.setStyle(TableStyle(t_styles))
    
    story.append(particulars_table)
    story.append(Spacer(1, 30))
    
    # --- FOOTER SECTION (Bank Details & Signature) ---
    bank_p = Paragraph(
        f"<b>Bank Account Details for Payment:</b><br/>"
        f"Account Name: {company_details.get('company_bank_account_name') or company_details.get('company_name', '')}<br/>"
        f"Bank Name: {company_details.get('company_bank_name', '')}<br/>"
        f"Account No: {company_details.get('company_bank_account', '')}<br/>"
        f"IFSC Code: {company_details.get('company_bank_ifsc', '')}<br/>"
        f"Branch: {company_details.get('company_bank_branch', '')}<br/>"
        f"Account Type: {company_details.get('company_bank_account_type', 'Current Account')}",
        normal_text
    )
    
    sign_p = Paragraph(
        f"For <b>{company_details.get('company_name', 'Company Name')}</b><br/><br/><br/><br/>"
        f"<b>{company_details.get('authorized_signatory', 'Nlesh Bajaj')}</b><br/>"
        f"{company_details.get('authorized_signatory_title', 'Authorized Signatory')}",
        ParagraphStyle('Sign', parent=normal_text, alignment=2) # Right align
    )
    
    footer_table = Table([[bank_p, sign_p]], colWidths=[300, 215])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    
    # Wrap in KeepTogether to ensure it doesn't break across pages
    story.append(KeepTogether(footer_table))
    
    # disputes disclaimer
    story.append(Spacer(1, 15))
    story.append(Paragraph("<font size='7' color='#718096'>All disputes restricted to Pune Jurisdiction</font>", normal_text))
    
    # Build Document
    doc.build(story)
    return True

def clean_label_val(val, prefix):
    if not val:
        return ""
    val_str = str(val).strip()
    import re
    # Remove prefix case insensitively
    pattern = re.compile(re.escape(prefix), re.IGNORECASE)
    cleaned = pattern.sub("", val_str).strip()
    # Remove leading colons or hyphens
    cleaned = re.sub(r'^[:\-\s]+', '', cleaned).strip()
    return cleaned

def find_master_sheet_name(sheetnames):
    priority = ["master file", "main sheet123", "main sheet", "masterdata"]
    for p in priority:
        for sname in sheetnames:
            if sname.strip().lower() == p:
                return sname
    for keyword in ["master", "main"]:
        for sname in sheetnames:
            if keyword in sname.lower():
                return sname
    return None

def get_master_sheet_mapping(ws):
    header_row = None
    cols = {}
    for r in range(1, 11):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(25, ws.max_column + 1))]
        normalized_vals = [str(v).strip().lower() if v is not None else "" for v in row_vals]
        if any(h in normalized_vals for h in ["client name", "name of client", "client_name", "name_of_client"]):
            header_row = r
            for col_idx, val in enumerate(row_vals, 1):
                if val:
                    normalized_h = str(val).strip().lower().replace(" ", "_")
                    cols[normalized_h] = col_idx
            break
    return header_row, cols

def get_mapped_col(cols_mapping, keys_list, fallback_idx=None):
    for k in keys_list:
        if k in cols_mapping:
            return cols_mapping[k]
    for k in keys_list:
        for mk in cols_mapping:
            if mk.startswith(k) or k in mk:
                return cols_mapping[mk]
    return fallback_idx

def parse_company_and_bank_details(ws_inv):
    company_details = {}
    company_details["company_name"] = str(ws_inv.cell(row=2, column=4).value or ws_inv.cell(row=1, column=4).value or "Cedrus Consultants Pvt Ltd").strip()
    company_details["company_division"] = str(ws_inv.cell(row=3, column=4).value or "").strip()
    company_details["company_address"] = str(ws_inv.cell(row=4, column=4).value or "").strip()
    
    for r in range(1, 16):
        for c in [4, 2]:
            val = ws_inv.cell(row=r, column=c).value
            if val:
                val_str = str(val)
                if "mobile" in val_str.lower() or "phone" in val_str.lower():
                    company_details["company_phone"] = clean_label_val(val, "Mobile Number")
                elif "pan" in val_str.lower():
                    company_details["company_pan"] = clean_label_val(val, "Pan Number")
                elif "gstin" in val_str.lower() or "gst" in val_str.lower():
                    company_details["company_gstin"] = clean_label_val(val, "GSTIN")

    if "company_phone" not in company_details: company_details["company_phone"] = ""
    if "company_pan" not in company_details: company_details["company_pan"] = ""
    if "company_gstin" not in company_details: company_details["company_gstin"] = ""

    for r in range(25, min(46, ws_inv.max_row + 1)):
        val = ws_inv.cell(row=r, column=2).value
        if val:
            val_str = str(val).lower()
            if "account name" in val_str or "acc name" in val_str:
                company_details["company_bank_account_name"] = clean_label_val(val, "Account Name")
            elif "bank account details" in val_str or "bank details" in val_str or "bank name" in val_str:
                company_details["company_bank_name"] = clean_label_val(val, "Bank Account Details")
                if not company_details["company_bank_name"]:
                    company_details["company_bank_name"] = clean_label_val(val, "Bank Details")
            elif "branch" in val_str:
                company_details["company_bank_branch"] = clean_label_val(val, "Bank Branch Address")
            elif "account type" in val_str or "acc type" in val_str:
                company_details["company_bank_account_type"] = clean_label_val(val, "Account Type")
            elif "account number" in val_str or "acc number" in val_str or "acc no" in val_str or "account no" in val_str:
                company_details["company_bank_account"] = clean_label_val(val, "Account Number")
            elif "ifsc" in val_str:
                company_details["company_bank_ifsc"] = clean_label_val(val, "IFSC Code")
                
        for c in [8, 7]:
            sig_val = ws_inv.cell(row=r, column=c).value
            if sig_val:
                sig_str = str(sig_val).strip()
                if sig_str.lower() in ["authorized signatory", "authorised signatory"]:
                    company_details["authorized_signatory_title"] = sig_str
                    for offset in [1, -1]:
                        name_val = ws_inv.cell(row=r + offset, column=c).value
                        if name_val and str(name_val).strip() and "signatory" not in str(name_val).lower():
                            company_details["authorized_signatory"] = str(name_val).strip()
                            break

    if "company_bank_account_name" not in company_details: company_details["company_bank_account_name"] = ""
    if "company_bank_name" not in company_details: company_details["company_bank_name"] = ""
    if "company_bank_branch" not in company_details: company_details["company_bank_branch"] = ""
    if "company_bank_account_type" not in company_details: company_details["company_bank_account_type"] = ""
    if "company_bank_account" not in company_details: company_details["company_bank_account"] = ""
    if "company_bank_ifsc" not in company_details: company_details["company_bank_ifsc"] = ""
    if "authorized_signatory" not in company_details: company_details["authorized_signatory"] = ""
    if "authorized_signatory_title" not in company_details: company_details["authorized_signatory_title"] = "Authorized Signatory"

    return company_details

def get_best_sheet_match(client_name, sheet_names):
    import re
    import difflib
    c_clean = re.sub(r'[^a-zA-Z0-9\s]', '', client_name).lower().strip()
    c_words = [w for w in c_clean.split() if w not in ('huf', 'and', 'trust', 'pvt', 'ltd', 'private', 'limited')]
    if not c_words:
        return None
    
    best_match = None
    max_score = 0
    
    for sname in sheet_names:
        sname_lower = sname.lower().strip()
        if sname_lower in ('master file', 'main sheet123', 'main sheet', 'masterdata', 'sheet1', 'invoice', 'invoice making mircro', 'new latest', 'sheet2', 'make invoice', 'update'):
            continue
        
        s_clean = re.sub(r'[^a-zA-Z0-9\s]', '', sname_lower)
        s_clean = re.sub(r'^\d+\s+', '', s_clean).strip()
        s_clean = re.sub(r'^va\d*\s+', '', s_clean).strip()
        
        s_words = [w for w in s_clean.split() if w not in ('huf', 'and', 'trust', 'pvt', 'ltd', 'private', 'limited')]
        if not s_words:
            continue
            
        first_word_matches = difflib.get_close_matches(c_words[0], s_words, n=1, cutoff=0.8)
        if not first_word_matches:
            continue
            
        fuzzy_matches = 0
        for cw in c_words:
            matches = difflib.get_close_matches(cw, s_words, n=1, cutoff=0.8)
            if matches:
                fuzzy_matches += 1
        score = fuzzy_matches
        
        c_joined = "".join(c_words)
        s_joined = "".join(s_words)
        if c_joined == s_joined:
            score += 5
        elif c_joined in s_joined or s_joined in c_joined:
            score += 2
            
        if score > max_score and score >= 1:
            max_score = score
            best_match = sname
            
    return best_match

def generate_invoice_pdf_via_excel(excel_path, client_key, pdf_path):
    if not EXCEL_AVAILABLE:
        raise RuntimeError("Microsoft Excel is not installed or pywin32 is not configured.")
        
    import pythoncom
    import win32com.client
    import os
    
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_excel = os.path.abspath(excel_path)
        abs_pdf = os.path.abspath(pdf_path)
        
        os.makedirs(os.path.dirname(abs_pdf), exist_ok=True)
        wb = excel.Workbooks.Open(abs_excel)
        
        ws = wb.Sheets("Invoice")
        ws.Range("B13").Value = client_key
        ws.Calculate()
        wb.Save()
        ws.ExportAsFixedFormat(0, abs_pdf)
        wb.Close(False)
        return True
    except Exception as e:
        print(f"Excel COM Invoice export failed: {e}")
        raise e
    finally:
        if wb:
            try:
                wb.Close(False)
            except:
                pass
        if excel:
            try:
                excel.Quit()
            except:
                pass
        pythoncom.CoUninitialize()

def parse_single_excel_sheet(excel_path):
    """
    Redesigned to dynamically parse master sheets like CCPL and Cedrus Finance.
    Returns (company_details, clients) where company_details may be None.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Error loading single Excel file {excel_path}: {e}")
        return None, []

    master_sheet = find_master_sheet_name(wb.sheetnames)
    if master_sheet:
        company_details = {}
        if "Invoice" in wb.sheetnames:
            ws_inv = wb["Invoice"]
            company_details = parse_company_and_bank_details(ws_inv)
        
        clients = []
        ws_master = wb[master_sheet]
        header_row, cols = get_master_sheet_mapping(ws_master)
        
        if header_row is None:
            wb.close()
            return None, []
            
        col_name = get_mapped_col(cols, ["client_name", "name_of_client"], 3)
        col_email = get_mapped_col(cols, ["client_email_id", "email_id", "email"], 9)
        col_invoice = get_mapped_col(cols, ["invoice_no", "invoice_number"], 5)
        col_rate = get_mapped_col(cols, ["fee_@", "rate"], 4)
        col_address = get_mapped_col(cols, ["address_1", "address"], 6)
        col_state = get_mapped_col(cols, ["state"], 7)
        col_particulars = get_mapped_col(cols, ["particular", "particulars"], 13)
        col_start = get_mapped_col(cols, ["from", "period_start"], 14)
        col_end = get_mapped_col(cols, ["to", "period_end"], 15)
        col_valuation = get_mapped_col(cols, ["value", "valuation", "value_of_shares", "aum"], 11)
        col_taxable = get_mapped_col(cols, ["taxable_amt", "taxable_amount"], 17)
        col_cgst = get_mapped_col(cols, ["cgst"], 18)
        col_sgst = get_mapped_col(cols, ["sgst"], 19)
        col_igst = get_mapped_col(cols, ["igst", "isgt"], 20)
        col_total = get_mapped_col(cols, ["total", "total_inv_amt", "total_amount"], 21)

        for r in range(header_row + 1, ws_master.max_row + 1):
            client_name = ws_master.cell(row=r, column=col_name).value
            if not client_name:
                continue
            
            email = str(ws_master.cell(row=r, column=col_email).value or "").strip()
            invoice_no = str(ws_master.cell(row=r, column=col_invoice).value or "").strip()
            
            rate_val = ws_master.cell(row=r, column=col_rate).value
            try:
                rate = float(rate_val) if rate_val is not None else 0.0
            except (ValueError, TypeError):
                rate = 0.0
                
            address = str(ws_master.cell(row=r, column=col_address).value or "").strip()
            state = str(ws_master.cell(row=r, column=col_state).value or "").strip()
            particulars = str(ws_master.cell(row=r, column=col_particulars).value or "").strip()
            
            val_start = ws_master.cell(row=r, column=col_start).value
            val_end = ws_master.cell(row=r, column=col_end).value
            
            try:
                valuation_val = ws_master.cell(row=r, column=col_valuation).value
                valuation = float(valuation_val) if valuation_val is not None else 0.0
            except (ValueError, TypeError):
                valuation = 0.0
                
            try:
                taxable_val = ws_master.cell(row=r, column=col_taxable).value
                taxable_amount = float(taxable_val) if taxable_val is not None else 0.0
            except (ValueError, TypeError):
                taxable_amount = 0.0
                
            try:
                cgst_val = ws_master.cell(row=r, column=col_cgst).value
                cgst = float(cgst_val) if cgst_val is not None else 0.0
            except (ValueError, TypeError):
                cgst = 0.0
                
            try:
                sgst_val = ws_master.cell(row=r, column=col_sgst).value
                sgst = float(sgst_val) if sgst_val is not None else 0.0
            except (ValueError, TypeError):
                sgst = 0.0
                
            try:
                igst_val = ws_master.cell(row=r, column=col_igst).value
                igst = float(igst_val) if igst_val is not None else 0.0
            except (ValueError, TypeError):
                igst = 0.0
                
            try:
                total_val = ws_master.cell(row=r, column=col_total).value
                total_amount = float(total_val) if total_val is not None else 0.0
            except (ValueError, TypeError):
                total_amount = 0.0
                
            if "#value" in state.lower() or not state:
                if igst > 0:
                    state = "Out of State"
                else:
                    state = "Maharashtra"
            
            matched_sheet = get_best_sheet_match(str(client_name), wb.sheetnames)
            
            clients.append({
                "client_name": str(client_name).strip(),
                "client_type": "Type 1",
                "state": state,
                "email": email,
                "cc_email": "",
                "address": address,
                "gstin": "",
                "valuation": valuation,
                "is_regular": total_amount > 0,
                "custom_message": "",
                "filename": excel_path,
                "status": "Pending",
                "matched_sheet": matched_sheet,
                "fee_amount": taxable_amount,
                "cgst": cgst,
                "sgst": sgst,
                "igst": igst,
                "total_amount": total_amount,
                "invoice_number": invoice_no,
                "rate": rate,
                "particulars": particulars,
                "period_start": str(val_start) if val_start else "",
                "period_end": str(val_end) if val_end else ""
            })
            
        wb.close()
        return company_details, clients

    # Fallback to standard sheet parser
    sheet = wb.active
    for name in wb.sheetnames:
        if name.strip().lower() in ["clients", "invoices", "master", "billing", "sheet1"]:
            sheet = wb[name]
            break

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return None, []

    # Find headers row
    header_row_idx = 0
    headers = None
    keywords = ["client", "name", "email", "valuation", "aum", "value", "state", "gst"]
    for idx, row in enumerate(rows):
        if any(isinstance(val, str) and any(kw in val.lower() for kw in keywords) for val in row if val is not None):
            headers = [str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in row]
            header_row_idx = idx
            break

    if headers is None:
        headers = [str(h).strip().lower().replace(" ", "_") if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        header_row_idx = 0

    clients = []
    
    for r_idx in range(header_row_idx + 1, len(rows)):
        row_vals = rows[r_idx]
        if not row_vals or not any(v is not None for v in row_vals):
            continue

        row_dict = {}
        for col_idx, h in enumerate(headers):
            if h and col_idx < len(row_vals):
                row_dict[h] = row_vals[col_idx]

        client_name = ""
        client_type = "Type 1"
        state = "Maharashtra"
        email = ""
        cc_email = ""
        address = ""
        gstin = ""
        valuation = 0.0
        is_regular = True
        custom_message = ""

        for k, v in row_dict.items():
            if v is None:
                continue
            k_lower = k.lower()
            if "client_name" in k_lower or "customer_name" in k_lower or k_lower in ["name", "client", "customer"]:
                client_name = str(v).strip()
            elif "client_type" in k_lower or k_lower in ["type", "classification"]:
                client_type = str(v).strip()
            elif k_lower in ["state", "supply", "place_of_supply", "supply_place"]:
                state = str(v).strip()
            elif k_lower in ["email", "email_address", "mail", "to"]:
                email = str(v).strip()
            elif k_lower in ["cc", "cc_email", "email_cc"]:
                cc_email = str(v).strip()
            elif "address" in k_lower:
                address = str(v).strip()
            elif "gstin" in k_lower or k_lower == "gst":
                gstin = str(v).strip()
            elif "valuation" in k_lower or "value" in k_lower or "portfolio" in k_lower or k_lower == "aum":
                try:
                    if isinstance(v, str):
                        v = v.replace("₹", "").replace("$", "").replace(",", "").strip()
                    valuation = float(v)
                except:
                    valuation = 0.0
            elif "regular" in k_lower or "active" in k_lower or "billable" in k_lower or k_lower == "status":
                val_str = str(v).strip().lower()
                if val_str in ["no", "false", "0", "inactive", "non-regular"]:
                    is_regular = False
                else:
                    is_regular = True
            elif "message" in k_lower or "body" in k_lower or "msg" in k_lower:
                custom_message = str(v).strip()

        if not client_name and (email or valuation > 0):
            client_name = f"Client Row {r_idx + 1}"

        if client_name:
            clients.append({
                "client_name": client_name,
                "client_type": client_type,
                "state": state,
                "email": email,
                "cc_email": cc_email,
                "address": address,
                "gstin": gstin,
                "valuation": valuation,
                "is_regular": is_regular,
                "custom_message": custom_message,
                "filename": excel_path,
                "status": "Pending"
            })

    wb.close()
    return None, clients

